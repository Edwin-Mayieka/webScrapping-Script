from selenium import webdriver
import time
import datetime
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import os
# Set the path to the Chrome binary
chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = '/usr/bin/chromium'

# excel file name
current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
filename='scholars_'+ current_time + '.xlsx'


# Set up the Chrome driver with the downloaded ChromeDriver executable
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

#= Navigate to the Schoology login page
driver.get('https://app.schoology.com/')

# Find the username, password, and school code fields and fill them
username_field = driver.find_element(By.ID, 'edit-mail')
password_field = driver.find_element(By.ID, 'edit-pass')
school_code_field = driver.find_element(By.ID, 'edit-school')
school_id_field=driver.find_element(By.ID, 'edit-school-nid')
school_code = '827110743'

username_field.send_keys('A00267623')
password_field.send_keys('090502')
time.sleep(3) 
password_field.click()
driver.execute_script("arguments[0].value = arguments[1];", school_id_field, school_code)
# Submit the login form
driver.find_element(By.ID, 'edit-submit').click()
driver.implicitly_wait(10)
driver.get('https://app.schoology.com/course/6667445896/members')

emails = []
phones=[]
names=[]
def loadData():
    wait = WebDriverWait(driver, 10)
    table = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'table[role="presentation"]')))
    ids=[]
    rows = table.find_elements(By.TAG_NAME, 'tr')
    for row in rows:
        user_id = row.get_attribute('id')
        ids.append(user_id)
        element = row.find_element(By.CSS_SELECTOR, 'td.user-name a.sExtlink-processed')
        name = element.text
        #last_name = element.find_element(By.TAG_NAME, 'b').text
        names.append(name)
        #print(name)

    for id in ids:
        if id:
           
            # Construct the user page URL using the extracted user ID
            user_page_url = f'https://app.schoology.com/user/{id}/info' 
            driver.get(user_page_url)
            # Find the email element and extract the email address
            try:
                email_element = driver.find_element(By.CSS_SELECTOR, 'a.sExtlink-processed.mailto')
                email = email_element.get_attribute('href').replace('mailto:', '')
                emails.append(email)
            except NoSuchElementException:
                    emails.append('')
                   
            try:
                phone_element = driver.find_element(By.XPATH, "//td/a[@class='sExtlink-processed']")
                phone = phone_element.get_attribute('href').replace('tel:', '')
                phones.append(phone)
            except NoSuchElementException:
                    phones.append('')
                          
    driver.get('https://app.schoology.com/course/6667445896/members')
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))

def nextPage(pages):
    wait = WebDriverWait(driver, 10)

    # Add initial wait for the page to load completely
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
        for page in range(pages):
            next_page = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
            if next_page:
                next_page.click()
                print(f"Clicked {page+1} times")

            # Delay for 5 seconds to allow the page to load
                time.sleep(5)

                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.next')))
    except:
        print("No more pages")
        print_data_to_excel(names,emails,phones, filename=filename)

def print_data_to_excel(names, emails, phones, filename):
    current_dir = os.getcwd()
    file_path = os.path.join(current_dir, filename)
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add headers
    sheet.cell(row=1, column=1).value = "Names"
    sheet.cell(row=1, column=2).value = "Emails"
    sheet.cell(row=1, column=3).value = "Phone"

    # Print data to respective columns
    for i in range(len(names)):
        sheet.cell(row=i+2, column=1).value = names[i]
        sheet.cell(row=i+2, column=2).value = emails[i]
        sheet.cell(row=i+2, column=3).value =phones[i]


    # Save the workbook
    workbook.save(file_path)
    print("Data printed to Excel successfully.")


def getData(startPage,pages):
    if startPage==1:
        for page in range(pages):
            if page == 0:
                loadData()
            else:
                nextPage(page)
                loadData()
    else:
        nextPage(startPage)
        for page in range(pages):
            if page == 0:
                loadData()
            else:
                nextPage(page)
                loadData()
    print_data_to_excel(names,emails,phones,filename=filename)
    # Quit
    driver.quit()


# Enter the startPage and the number of pages that you want to go through
#pages=totalPages/totalItems-per-page ie for 600pages and 30 items per page = 600/30= 20 pages
getData(1,1)




